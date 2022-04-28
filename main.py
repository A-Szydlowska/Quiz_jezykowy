from openpyxl import load_workbook
from openpyxl import Workbook
import random

rand = []
number = 0
category = 0
print("Witaj w programie do testu ze słówek!")
print("Jeśli chcesz skorzystać z domyślnego zestawu słówek wybierz 1 a następnie wpisz nazwę 'test.xlsx'\n")
while 1:
    try:
        print("Proszę wybierz co chcesz zrobić:\n")
        choice = int(
            input("1-utwórz test\n2-popraw istniejący test\n3-pokaż wyniki\n4-wczytaj baze słówek\n5-zakończ\n\n"))
    except ValueError:
        print("Wprowadziłeś niepoprawną liczbę\n")
        continue
    else:
        if choice < 1 or choice > 5:
            print("Wprowadziłeś niepoprawną liczbę\n")
            continue
        if choice == 1:
            choose = input("Podaj nazwę pliku z bazą słówek (np. test.xlsx)\n")
            wb = load_workbook(choose)
            sheet = wb['Baza']

            name = input("Podaj swoje imię\n")
            surname = input("Podaj swoje nazwisko\n")
            while 1:
                i = 2
                j = 1
                for row in sheet.iter_rows(min_row=1, min_col=2, max_row=1, max_col=6):
                    for cell in row:
                        print(j, sheet.cell(row=1, column=i).value)
                        i += 2
                        j += 1
                try:
                    category = int(input("Podaj numer kategorii od 1 do 5:\n"))
                except ValueError:
                    print("Wprowadziłeś niepoprawną liczbę\n")
                    continue
                else:
                    if category < 1 or category > 5:
                        print("Wprowadziłeś niepoprawną liczbę")
                        continue
                    else:
                        break
            while 1:
                try:
                    number = int(input("Podaj liczbę słówek do wylosowania w teście (1-20)\n"))
                except ValueError:
                    print("Wprowadziłeś niepoprawną liczbę\n")
                    continue
                else:
                    if number < 1 or number > 20:
                        print("Wprowadziłeś niepoprawną liczbę")
                        continue
                    else:
                        break
            result = 0
            mistakes = []
            answers = []

            for k in range(0, number):
                while 1:
                    word = random.randint(2, 21)
                    if word in rand:
                        continue
                    else:
                        rand.append(word)
                        break
                print(sheet.cell(row=word, column=category * 2).value)
                print("\n")
                answer = input()
                answers.append(sheet.cell(row=word, column=category * 2).value)
                answers.append(answer)
                if answer == sheet.cell(row=word, column=category * 2 + 1).value:
                    result += 1
                else:
                    mistakes.append(sheet.cell(row=word, column=category * 2).value)
                    mistakes.append(sheet.cell(row=word, column=category * 2 + 1).value)
                    result -= 1
            print(
                "Test zakończył się. Aby poznać swój wynik wprowadź nazwę pliku xlsx do której ma zostać wgrany "
                "raport (np. test.xlsx)\n")
            work = input()  # tworzenie arkusza
            wb = Workbook()
            wb.save(filename=work)
            wb = load_workbook(work)
            sheet = wb['Sheet']
            sheet.title = 'Raport'
            m = 1
            n = 0
            for row in sheet.iter_cols(min_row=1, min_col=1, max_row=int(len(answers) / 2), max_col=1):
                for cell in row:
                    if n > len(answers) - 1:
                        break
                    sheet.cell(row=m, column=1).value = answers[n]
                    sheet.cell(row=m, column=2).value = answers[n + 1]
                    m += 1
                    n += 2
            m = n
            n = 0
            if len(mistakes) != 0:
                for row in sheet.iter_cols(min_row=5, min_col=1, max_row=5 + int(len(mistakes) / 2), max_col=1):
                    for cell in row:
                        if n > len(mistakes) - 1:
                            break
                        sheet.cell(row=m, column=1).value = mistakes[n]
                        sheet.cell(row=m, column=2).value = mistakes[n + 1]
                        m += 1
                        n += 2
            sheet.cell(row=m, column=1).value = "Ilość zdobytych punktów:"
            sheet.cell(row=m, column=2).value = result
            m += 1
            sheet.cell(row=m, column=1).value = "Ocena:"
            if result / number < 0.5:
                sheet.cell(row=m, column=2).value = 2
            if 0.5 <= result / number < 0.7:
                sheet.cell(row=m, column=2).value = 3
            if 0.75 <= result / number < 0.9:
                sheet.cell(row=m, column=2).value = 4
            if result / number >= 0.9:
                sheet.cell(row=m, column=2).value = 5

            wb.create_sheet("Konfiguracja")  # tworzenie arkusza konfiguracyjego
            sheet = wb['Konfiguracja']
            sheet.cell(row=1, column=1).value = "Imię"
            sheet.cell(row=1, column=2).value = name
            sheet.cell(row=2, column=1).value = "Nazwisko"
            sheet.cell(row=2, column=2).value = surname
            sheet.cell(row=3, column=1).value = "Numer kategorii"
            sheet.cell(row=3, column=2).value = category
            sheet.cell(row=4, column=1).value = "Ilość słówek"
            sheet.cell(row=4, column=2).value = number

            wb.save(work)

        if choice == 2:
            report = input("Podaj nazwę pliku z raportem (np. test.xlsx)\n")
            wb = load_workbook(report)
            sheet = wb['Konfiguracja']
            category = sheet.cell(row=3, column=2).value
            number = sheet.cell(row=4, column=2).value

            choose = input("Podaj nazwę pliku z bazą słówek (np. test.xlsx)\n")
            wb = load_workbook(choose)
            sheet = wb['Baza']

            result = 0
            mistakes = []
            answers = []
            for k in range(0, number):
                while 1:
                    word = random.randint(2, 21)
                    if word in rand:
                        continue
                    else:
                        rand.append(word)
                        break
                print(sheet.cell(row=word, column=category * 2).value)
                print("\n")
                answer = input()
                answers.append(sheet.cell(row=word, column=category * 2).value)
                answers.append(answer)
                if answer == sheet.cell(row=word, column=category * 2 + 1).value:
                    result += 1
                else:
                    mistakes.append(sheet.cell(row=word, column=category * 2).value)
                    mistakes.append(sheet.cell(row=word, column=category * 2 + 1).value)
                    result -= 1

            wb = load_workbook(report)
            wb.create_sheet("Poprawa")
            sheet = wb['Poprawa']
            m = 1
            n = 0
            for row in sheet.iter_cols(min_row=1, min_col=1, max_row=int(len(answers) / 2), max_col=1):
                for cell in row:
                    if n > len(answers) - 1:
                        break
                    sheet.cell(row=m, column=1).value = answers[n]
                    sheet.cell(row=m, column=2).value = answers[n + 1]
                    m += 1
                    n += 2
            m = n
            n = 0
            if len(mistakes) != 0:
                for row in sheet.iter_cols(min_row=5, min_col=1, max_row=5 + int(len(mistakes) / 2), max_col=1):
                    for cell in row:
                        if n > len(mistakes) - 1:
                            break
                        sheet.cell(row=m, column=1).value = mistakes[n]
                        sheet.cell(row=m, column=2).value = mistakes[n + 1]
                        m += 1
                        n += 2
            sheet.cell(row=m, column=1).value = "Ocena:"
            if result / number < 0.5:
                sheet.cell(row=m, column=2).value = 2
            if 0.5 <= result / number < 0.7:
                sheet.cell(row=m, column=2).value = 3
            if 0.75 <= result / number < 0.9:
                sheet.cell(row=m, column=2).value = 4
            if result / number >= 0.9:
                sheet.cell(row=m, column=2).value = 5

            wb.save(report)
            print("Test został zakończony pomyślnie\n")

        if choice == 3:
            report = input("Podaj nazwę pliku z raportem (np. test.xlsx)\n")
            wb = load_workbook(report)
            while 1:
                choice = input(
                    "Czy chcesz wyświetlić raport poprawy testu? (t/n)\n")
                if choice == 't':
                    sheet = wb['Poprawa']
                    i = 1
                    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=50, max_col=2):
                        for cell in row:
                            if sheet.cell(row=i, column=1).value is not None:
                                print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)
                            i += 1
                    break
                if choice == 'n':
                    sheet = wb['Raport']
                    i = 1
                    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=50, max_col=2):
                        for cell in row:
                            if sheet.cell(row=i, column=1).value is not None:
                                print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)
                            i += 1
                    break
                if choice != 't' and choice != 'n':
                    print("Niepoprawny znak, spróbuj jeszcze raz.\n")
                    continue
        if choice == 4:
            work = input("Podaj nazwę pliku do którego ma być wczytana baza słówek (np. test.xlsx)\n")
            wb = Workbook()
            wb.save(filename=work)
            wb = load_workbook(work)
            sheet = wb['Sheet']
            sheet.title = 'Baza'
            i = 2
            j = 1
            for row in sheet.iter_cols(min_row=2, min_col=1, max_row=21, max_col=1):
                for cell in row:
                    sheet.cell(row=i, column=1).value = j
                    j += 1
                    i += 1
            i = 2
            for k in range(0, 5):
                cat = input("Wprowadź nazwę kategorii\n")
                j = 1
                sheet.cell(row=j, column=i).value = cat
                for o in range(0, 20):
                    pol = input("Podaj słówko polskie\n")
                    sheet.cell(row=o + 2, column=i).value = pol
                    eng = input("Podaj słówko angielskie\n")
                    sheet.cell(row=o + 2, column=i + 1).value = eng
                i += 2
            wb.save(work)
        if choice == 5:
            break
print("Dziękuję za skorzystanie z mojego programu!")
